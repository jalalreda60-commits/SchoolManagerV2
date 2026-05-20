import os
from datetime import datetime
from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QComboBox, QSpinBox, QMessageBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QGridLayout, QSizePolicy)
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor, QFont

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MONTHS = ['Janvier','Février','Mars','Avril','Mai','Juin',
          'Juillet','Août','Septembre','Octobre','Novembre','Décembre']


class ReportCard(QFrame):
    def __init__(self, title, desc, icon, color, callback):
        super().__init__()
        self.setObjectName("card")
        self.setMinimumHeight(140)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.setStyleSheet(f"""
            QFrame#card{{
                background:#1e293b; border:1px solid #1e3a5f;
                border-top:3px solid {color}; border-radius:14px;
            }}
            QFrame#card:hover{{background:#162032; border-color:{color};}}
        """)
        lay = QVBoxLayout(self); lay.setContentsMargins(20,18,20,18); lay.setSpacing(8)
        top = QHBoxLayout()
        ic = QLabel(icon); ic.setStyleSheet(f"background:transparent; font-size:24px; color:{color};")
        top.addWidget(ic); top.addStretch()
        lay.addLayout(top)
        t = QLabel(title); t.setStyleSheet(f"background:transparent; color:{color}; font-size:14px; font-weight:700;")
        lay.addWidget(t)
        d = QLabel(desc); d.setStyleSheet("background:transparent; color:#94a3b8; font-size:11px;")
        d.setWordWrap(True); lay.addWidget(d)
        btn = QPushButton("📥  Générer PDF")
        btn.setStyleSheet(f"""
            QPushButton{{background:rgba({self._rgb(color)},0.12); color:{color};
                        border:1px solid rgba({self._rgb(color)},0.3); border-radius:8px;
                        padding:7px 14px; font-weight:600; font-size:12px;}}
            QPushButton:hover{{background:rgba({self._rgb(color)},0.25);}}
        """)
        btn.clicked.connect(callback)
        lay.addWidget(btn)

    def _rgb(self, h):
        h = h.lstrip('#')
        return f"{int(h[0:2],16)},{int(h[2:4],16)},{int(h[4:6],16)}"


class ReportsWidget(QWidget):
    def __init__(self, session, current_user):
        super().__init__()
        self.session = session; self.current_user = current_user
        self.setStyleSheet("background:transparent;")
        self._build()

    def _build(self):
        from ui.widgets import page_header, filter_card, SummaryPill
        root = QVBoxLayout(self); root.setContentsMargins(24,20,24,20); root.setSpacing(16)

        hdr, _ = page_header("Rapports & Exports",
            "Générez des rapports PDF professionnels et exportez en Excel.")
        root.addWidget(hdr)

        # Report cards grid
        grid = QGridLayout(); grid.setSpacing(14)
        reports = [
            ("Liste Élèves","Dossiers complets de tous les élèves actifs","🎓","#3b82f6",self._rpt_students),
            ("Rapport Financier","Revenus, dépenses et bénéfice mensuel/annuel","💰","#22c55e",self._rpt_financial),
            ("Élèves Impayés","Liste des élèves avec des paiements en retard","⚠","#f59e0b",self._rpt_unpaid),
            ("Rapport Employés","Effectif, postes et masse salariale","👥","#8b5cf6",self._rpt_employees),
        ]
        for i, (t, d, ic, col, fn) in enumerate(reports):
            grid.addWidget(ReportCard(t, d, ic, col, fn), 0, i)
        gw = QFrame(); gw.setLayout(grid); gw.setStyleSheet("background:transparent;")
        root.addWidget(gw)

        # Filter + Excel
        fc_lay = QHBoxLayout(); fc_lay.setSpacing(12)
        ml = QLabel("Mois:"); ml.setStyleSheet("color:#94a3b8; font-size:12px; background:transparent;")
        self._month_f = QComboBox(); self._month_f.setFixedHeight(38); self._month_f.setFixedWidth(130)
        for m in MONTHS: self._month_f.addItem(m)
        self._month_f.setCurrentIndex(datetime.now().month-1)
        self._month_f.currentIndexChanged.connect(self._load_preview)
        yl = QLabel("Année:"); yl.setStyleSheet("color:#94a3b8; font-size:12px; background:transparent;")
        self._year_f = QSpinBox(); self._year_f.setRange(2020,2050)
        self._year_f.setValue(datetime.now().year); self._year_f.setFixedHeight(38); self._year_f.setFixedWidth(80)
        self._year_f.valueChanged.connect(self._load_preview)
        excel_btn = QPushButton("📊  Exporter Excel")
        excel_btn.setObjectName("successBtn"); excel_btn.setFixedHeight(38)
        excel_btn.clicked.connect(self._export_excel)
        fc_lay.addWidget(ml); fc_lay.addWidget(self._month_f)
        fc_lay.addWidget(yl); fc_lay.addWidget(self._year_f)
        fc_lay.addStretch(); fc_lay.addWidget(excel_btn)
        ff = QFrame(); ff.setLayout(fc_lay); ff.setStyleSheet("background:transparent;")
        root.addWidget(ff)

        # Preview label
        prev_lbl = QLabel("📋  Aperçu — Paiements du mois sélectionné")
        prev_lbl.setStyleSheet("color:#f8fafc; font-size:14px; font-weight:600; background:transparent;")
        root.addWidget(prev_lbl)

        self._preview = QTableWidget()
        self._preview.setColumnCount(5)
        self._preview.setHorizontalHeaderLabels(["Reçu","Élève","Type","Montant","Date"])
        self._preview.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self._preview.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._preview.verticalHeader().setVisible(False)
        self._preview.setShowGrid(False)
        self._preview.setAlternatingRowColors(True)
        root.addWidget(self._preview)
        self._load_preview()

    def _load_preview(self):
        from models.database import Payment
        month = self._month_f.currentIndex()+1
        year  = self._year_f.value()
        pays  = self.session.query(Payment).filter_by(month=month, year=year).all()
        tlabels = {'monthly':'Mensualité','insurance':'Assurance','transport':'Transport'}
        self._preview.setRowCount(0)
        for p in pays:
            r = self._preview.rowCount(); self._preview.insertRow(r)
            st = p.student
            for c, v in enumerate([p.receipt_number or "—",
                f"{st.first_name} {st.last_name}" if st else "—",
                tlabels.get(p.payment_type, p.payment_type),
                f"{p.amount:,.2f} MAD",
                p.payment_date.strftime("%d/%m/%Y") if p.payment_date else "—"]):
                item = QTableWidgetItem(v); item.setTextAlignment(Qt.AlignCenter)
                if c == 3: item.setForeground(QColor("#22c55e"))
                self._preview.setItem(r, c, item)
            self._preview.setRowHeight(r, 38)

    # ── PDF generators ─────────────────────────────────────────────────────────
    def _open(self, path):
        import subprocess, platform
        try:
            if platform.system()=='Windows': os.startfile(path)
            elif platform.system()=='Darwin': subprocess.run(['open',path])
            else: subprocess.run(['xdg-open',path])
        except: pass

    def _rpt_students(self):
        from models.database import Student
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RI
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors; from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        rdir = os.path.join(BASE_DIR,'reports'); os.makedirs(rdir,exist_ok=True)
        fp = os.path.join(rdir,f"students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        doc = SimpleDocTemplate(fp,pagesize=landscape(A4),
            rightMargin=15*mm,leftMargin=15*mm,topMargin=15*mm,bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        ts = ParagraphStyle('t',fontSize=16,fontName='Helvetica-Bold',
            textColor=colors.HexColor('#3b82f6'),alignment=TA_CENTER)
        story = [Paragraph("LISTE DES ÉLÈVES — LE SCHÉMA", ts),
                 Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                    ParagraphStyle('s',fontSize=9,textColor=colors.grey,alignment=TA_CENTER)),
                 Spacer(1,8*mm)]
        students = self.session.query(Student).filter_by(is_active=True).all()
        data = [["Code","Nom","Prénom","Classe","Parent","Téléphone","Transport","Assurance"]]
        for s in students:
            data.append([s.student_code or "",s.last_name,s.first_name,
                s.class_.name if s.class_ else "—",
                s.parent_name or "—",s.parent_phone or "—",
                "✓" if s.has_transport else "✗",
                "✓" if s.insurance_paid else "✗"])
        tbl = Table(data,repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#3b82f6')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,-1),8),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f0f9ff')]),
            ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#e2e8f0')),
            ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
        ]))
        story.append(tbl)
        doc.build(story)
        self._open(fp)
        QMessageBox.information(self,"✅","Rapport élèves généré!")

    def _rpt_financial(self):
        from models.database import Payment, Expense
        from sqlalchemy import func
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors; from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_CENTER

        year = self._year_f.value()
        rdir = os.path.join(BASE_DIR,'reports'); os.makedirs(rdir,exist_ok=True)
        fp = os.path.join(rdir,f"financial_{year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        doc = SimpleDocTemplate(fp,pagesize=A4,
            rightMargin=20*mm,leftMargin=20*mm,topMargin=20*mm,bottomMargin=20*mm)
        ts = ParagraphStyle('t',fontSize=18,fontName='Helvetica-Bold',
            textColor=colors.HexColor('#22c55e'),alignment=TA_CENTER)
        story = [Paragraph(f"RAPPORT FINANCIER {year}", ts), Spacer(1,8*mm)]
        data = [["Mois","Recettes","Dépenses","Bénéfice"]]
        tot_r = tot_e = 0
        for m in range(1,13):
            r = self.session.query(func.sum(Payment.amount)).filter(
                Payment.month==m,Payment.year==year).scalar() or 0
            e = self.session.query(func.sum(Expense.amount)).filter(
                func.strftime('%m',Expense.expense_date)==f"{m:02d}",
                func.strftime('%Y',Expense.expense_date)==str(year)).scalar() or 0
            data.append([MONTHS[m-1],f"{r:,.2f} MAD",f"{e:,.2f} MAD",f"{r-e:,.2f} MAD"])
            tot_r += r; tot_e += e
        data.append(["TOTAL",f"{tot_r:,.2f} MAD",f"{tot_e:,.2f} MAD",f"{tot_r-tot_e:,.2f} MAD"])
        tbl = Table(data,repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#22c55e')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),
            ('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#0f172a')),
            ('TEXTCOLOR',(0,-1),(-1,-1),colors.white),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('ROWBACKGROUNDS',(0,1),(-1,-2),[colors.white,colors.HexColor('#f0fdf4')]),
            ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#e2e8f0')),
            ('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8),
        ]))
        story.append(tbl)
        doc.build(story); self._open(fp)
        QMessageBox.information(self,"✅","Rapport financier généré!")

    def _rpt_unpaid(self):
        from models.database import Student, Payment
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors; from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_CENTER

        month = self._month_f.currentIndex()+1; year = self._year_f.value()
        paid_ids = [p[0] for p in self.session.query(Payment.student_id).filter(
            Payment.month==month,Payment.year==year,Payment.payment_type=='monthly').all()]
        unpaid = self.session.query(Student).filter(
            Student.is_active==True,~Student.id.in_(paid_ids)).all()
        rdir = os.path.join(BASE_DIR,'reports'); os.makedirs(rdir,exist_ok=True)
        fp = os.path.join(rdir,f"unpaid_{year}_{month:02d}.pdf")
        doc = SimpleDocTemplate(fp,pagesize=A4,
            rightMargin=20*mm,leftMargin=20*mm,topMargin=20*mm,bottomMargin=20*mm)
        ts = ParagraphStyle('t',fontSize=16,fontName='Helvetica-Bold',
            textColor=colors.HexColor('#ef4444'),alignment=TA_CENTER)
        story = [Paragraph(f"ÉLÈVES IMPAYÉS — {MONTHS[month-1].upper()} {year}", ts),
                 Spacer(1,8*mm)]
        if not unpaid:
            story.append(Paragraph("✅ Tous les élèves sont à jour!",
                ParagraphStyle('ok',fontSize=14,textColor=colors.green,alignment=TA_CENTER)))
        else:
            data = [["Nom","Prénom","Classe","Frais","Parent","Téléphone"]]
            total = 0
            for s in unpaid:
                data.append([s.last_name,s.first_name,
                    s.class_.name if s.class_ else "—",
                    f"{s.monthly_fee:,.2f} MAD",
                    s.parent_name or "—",s.parent_phone or "—"])
                total += s.monthly_fee or 0
            tbl = Table(data,repeatRows=1)
            tbl.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#ef4444')),
                ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('ALIGN',(0,0),(-1,-1),'CENTER'),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#fef2f2')]),
                ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#e2e8f0')),
                ('TOPPADDING',(0,0),(-1,-1),7),('BOTTOMPADDING',(0,0),(-1,-1),7),
            ]))
            story.append(tbl)
            story.append(Spacer(1,5*mm))
            story.append(Paragraph(f"Total impayé: {total:,.2f} MAD | {len(unpaid)} élève(s)",
                ParagraphStyle('s',fontSize=11,fontName='Helvetica-Bold',
                    textColor=colors.HexColor('#ef4444'),alignment=TA_CENTER)))
        doc.build(story); self._open(fp)
        QMessageBox.information(self,"✅","Rapport impayés généré!")

    def _rpt_employees(self):
        from models.database import Employee
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors; from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_CENTER

        rdir = os.path.join(BASE_DIR,'reports'); os.makedirs(rdir,exist_ok=True)
        fp = os.path.join(rdir,f"employees_{datetime.now().strftime('%Y%m%d')}.pdf")
        doc = SimpleDocTemplate(fp,pagesize=A4,
            rightMargin=20*mm,leftMargin=20*mm,topMargin=20*mm,bottomMargin=20*mm)
        ts = ParagraphStyle('t',fontSize=16,fontName='Helvetica-Bold',
            textColor=colors.HexColor('#8b5cf6'),alignment=TA_CENTER)
        story = [Paragraph("RAPPORT DES EMPLOYÉS — LE SCHÉMA", ts),Spacer(1,8*mm)]
        emps = self.session.query(Employee).filter_by(is_active=True).all()
        tmap = {'teacher':'Enseignant','staff':'Personnel','driver':'Chauffeur'}
        data = [["Nom","Prénom","Type","Matière","Téléphone","Salaire"]]
        total = 0
        for e in emps:
            data.append([e.last_name,e.first_name,
                tmap.get(e.employee_type,e.employee_type),
                e.subject or "—",e.phone or "—",f"{e.base_salary:,.0f} MAD"])
            total += e.base_salary or 0
        tbl = Table(data,repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#8b5cf6')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#faf5ff')]),
            ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#e2e8f0')),
            ('TOPPADDING',(0,0),(-1,-1),7),('BOTTOMPADDING',(0,0),(-1,-1),7),
        ]))
        story.append(tbl)
        story.append(Spacer(1,5*mm))
        story.append(Paragraph(f"Masse salariale: {total:,.0f} MAD/mois | {len(emps)} employé(s)",
            ParagraphStyle('s',fontSize=11,textColor=colors.HexColor('#8b5cf6'),alignment=TA_CENTER)))
        doc.build(story); self._open(fp)
        QMessageBox.information(self,"✅","Rapport employés généré!")

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from models.database import Student, Payment
            from PySide6.QtWidgets import QFileDialog

            rdir = os.path.join(BASE_DIR,'reports'); os.makedirs(rdir,exist_ok=True)
            default = os.path.join(rdir,f"export_{datetime.now().strftime('%Y%m%d')}.xlsx")
            fp, _ = QFileDialog.getSaveFileName(self,"Enregistrer",default,"Excel (*.xlsx)")
            if not fp: return

            wb = openpyxl.Workbook()
            BLUE = "FF3b82f6"; WHITE = "FFFFFFFF"; LIGHT = "FFeff6ff"

            # Students sheet
            ws = wb.active; ws.title = "Élèves"
            headers = ["Code","Nom","Prénom","Classe","Parent","Téléphone","Frais","Transport","Assurance"]
            for ci, h in enumerate(headers,1):
                cell = ws.cell(row=1,column=ci,value=h)
                cell.font = Font(bold=True,color=WHITE)
                cell.fill = PatternFill("solid",fgColor=BLUE)
                cell.alignment = Alignment(horizontal='center')
            for ri, s in enumerate(self.session.query(Student).filter_by(is_active=True).all(),2):
                for ci, v in enumerate([s.student_code,s.last_name,s.first_name,
                    s.class_.name if s.class_ else "",s.parent_name,s.parent_phone,
                    s.monthly_fee,"Oui" if s.has_transport else "Non",
                    "Oui" if s.insurance_paid else "Non"],1):
                    cell = ws.cell(row=ri,column=ci,value=v)
                    cell.alignment = Alignment(horizontal='center')
                    if ri%2==0: cell.fill = PatternFill("solid",fgColor=LIGHT)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col)+4, 28)

            # Payments sheet
            ws2 = wb.create_sheet("Paiements")
            GREEN = "FF22c55e"; LIGHT2 = "FFf0fdf4"
            h2 = ["Reçu","Élève","Type","Mois","Année","Montant","Date"]
            for ci, h in enumerate(h2,1):
                cell = ws2.cell(row=1,column=ci,value=h)
                cell.font = Font(bold=True,color=WHITE)
                cell.fill = PatternFill("solid",fgColor=GREEN)
                cell.alignment = Alignment(horizontal='center')
            tmap = {'monthly':'Mensualité','insurance':'Assurance','transport':'Transport'}
            for ri, p in enumerate(self.session.query(Payment).order_by(
                    Payment.payment_date.desc()).limit(1000).all(),2):
                st = p.student
                for ci, v in enumerate([p.receipt_number,
                    f"{st.first_name} {st.last_name}" if st else "",
                    tmap.get(p.payment_type,p.payment_type),
                    MONTHS[p.month-1] if p.month and 1<=p.month<=12 else "",
                    str(p.year or ""),p.amount,
                    p.payment_date.strftime("%d/%m/%Y") if p.payment_date else ""],1):
                    cell = ws2.cell(row=ri,column=ci,value=v)
                    cell.alignment = Alignment(horizontal='center')
                    if ri%2==0: cell.fill = PatternFill("solid",fgColor=LIGHT2)
            for col in ws2.columns:
                ws2.column_dimensions[col[0].column_letter].width = 18

            wb.save(fp); self._open(fp)
            QMessageBox.information(self,"✅",f"Export Excel généré!")
        except Exception as e:
            QMessageBox.warning(self,"Erreur",f"Erreur export:\n{e}")
